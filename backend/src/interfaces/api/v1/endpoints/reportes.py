"""Exportación de reportes del SGC en formatos descargables.

- ``GET /reportes/excel`` -> Reporte Excel RA-105 (diario / semanal / mensual)
  consolidando las incidencias y sus casos de prueba desde todas las
  evaluaciones/trabajos.
"""
import io
from datetime import date, datetime, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.core.dependencies import get_current_user
from src.infrastructure.database.connection import get_db
from src.infrastructure.database.models.trabajo_model import (
    IncidenciaModel,
    EvaluacionModel,
    TrabajoModel,
)

router = APIRouter()

# Columnas definidas en la plantilla RA-105 solicitadas para el reporte.
COLUMNAS = [
    "N° Incidencia GLPI",
    "Descripción de la Incidencia",
    "Fecha de Registro de la Incidencia",
    "Prioridad",
    "Analista",
    "Tipo de Incidencia",
    "Base de Datos",
    "Versión",
    "Módulo",
    "Estado",
    "Observaciones",
    "Interacción",
    "N° Informe de Pruebas",
]


def _rango_periodo(periodo: str) -> tuple[datetime, datetime]:
    hoy = datetime.now()
    if periodo == "daily":
        inicio = hoy.replace(hour=0, minute=0, second=0, microsecond=0)
    elif periodo == "weekly":
        inicio = hoy - timedelta(days=7)
    elif periodo == "monthly":
        inicio = hoy.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    else:
        raise HTTPException(status_code=422, detail="Periodo inválido. Use daily, weekly o monthly.")
    return inicio, hoy


def _celda_fecha(valor):
    if isinstance(valor, (datetime, date)):
        return valor.strftime("%Y-%m-%d %H:%M")
    return valor or ""


async def _recolectar_filas(db: AsyncSession, inicio: datetime, fin: datetime) -> list[dict]:
    stmt = (
        select(IncidenciaModel)
        .where(IncidenciaModel.created_at >= inicio, IncidenciaModel.created_at <= fin)
        .options(
            selectinload(IncidenciaModel.evaluacion).selectinload(EvaluacionModel.trabajo),
        )
    )
    result = (await db.execute(stmt)).scalars().all()
    filas = []
    for i in result:
        e = i.evaluacion
        t = e.trabajo
        filas.append({
            "glpi": i.codigo or i.correlativo,
            "descripcion": i.descripcion,
            "fecha": _celda_fecha(i.created_at or e.fecha_asignacion),
            "prioridad": i.prioridad,
            "analista": e.analista,
            "tipo": i.tipo_error,
            "base_datos": i.base_datos,
            "version": i.version,
            "modulo": t.tipo_atencion,
            "estado": e.estado,
            "observaciones": "",
            "interaccion": "",
            "informe": "",
        })
    return filas


@router.get("/excel")
async def exportar_excel_ra105(
    periodo: str = Query("daily", pattern="^(daily|weekly|monthly)$"),
    current_user = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    inicio, fin = _rango_periodo(periodo)
    filas = await _recolectar_filas(db, inicio, fin)

    wb = Workbook()
    ws = wb.active
    ws.title = "RA-105"

    cabecera_fill = PatternFill("solid", fgColor="1a3a6b")
    cabecera_font = Font(bold=True, color="FFFFFF", size=11)
    borde = Border(*[Side(style="thin", color="999999")] * 4)

    for c, titulo in enumerate(COLUMNAS, start=1):
        cell = ws.cell(row=1, column=c, value=titulo)
        cell.fill = cabecera_fill
        cell.font = cabecera_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borde
    ws.row_dimensions[1].height = 30

    for r, fila in enumerate(filas, start=2):
        for c, clave in enumerate(
            # 01_glpi, 02_descripcion, 03_fecha, 04_prioridad, 05_analista,
            # 06_tipo, 07_base_datos, 08_version, 09_modulo, 10_estado,
            # 11_observaciones, 12_interaccion, 13_informe
            [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12],
            start=1,
        ):
            keys = ["glpi", "descripcion", "fecha", "prioridad", "analista", "tipo",
                    "base_datos", "version", "modulo", "estado", "observaciones",
                    "interaccion", "informe"]
            cell = ws.cell(row=r, column=c, value=fila[keys[clave]])
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = borde

    anchos = [18, 45, 20, 12, 22, 18, 16, 12, 22, 20, 30, 12, 18]
    for c, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(c)].width = ancho

    ws.freeze_panes = "A2"

    bufer = io.BytesIO()
    wb.save(bufer)
    bufer.seek(0)

    nombre = f"RA105_{periodo}_{date.today().isoformat()}.xlsx"
    return Response(
        content=bufer.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )
